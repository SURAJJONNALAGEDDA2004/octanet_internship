import pwinput
from openpyxl import load_workbook

excel_file = "atm_database.xlsx"
transaction_file = "transaction_history.xlsx"

class User:
    def __init__(self, user_id, pin, balance):
        self.user_id = user_id
        self.pin = pin
        self.balance = int(balance)

def load_users_from_excel():
    workbook = load_workbook(excel_file)
    sheet = workbook.active
    users = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        users.append(User(row[0], row[1], row[2]))
    return users, workbook

def update_balance_in_excel(users, workbook):
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_col=3, max_row=sheet.max_row):
        for user in users:
            if user.user_id == row[0].value:
                row[2].value = user.balance
                break
    workbook.save(excel_file)

def update_transaction_history(hjk):
    wb = load_workbook(transaction_file)
    sheet = wb.active
    sheet.append([hjk])
    wb.save(transaction_file)

def usage(user_id, pin, users, workbook):
    u = int(user_id)
    v = int(pin)
    for user in users:
        if user.user_id == u and user.pin == v:
            transactions = []
            while True:
                print("Welcome,what service would you like to use")
                print("Press 1 for Withdraw")
                print("Press 2 for Deposit")
                print("Press 3 for Transfer")
                print("Press 4 for transaction history")
                print("Press 5 to quit")
                print("Press 6 to check balance")
                ch=int(input("Enter choice: "))
                if(ch==1):
                    a=(input("Enter amount to withdraw: "))
                    if(a.isdigit()):
                        if(int(a)<=user.balance):
                            user.balance=int(user.balance)-int(a)
                            transactions.append((f"${a} has been withdrawn from your account with user id {u}."))
                            hjk= f"${a} has been withdrawn from your account with user id {u}."
                            print("Withdrawl complete,Remaining balance")
                            print(user.balance)
                            print("\n")
                            update_balance_in_excel(users, workbook)
                            update_transaction_history(hjk)
                        else:
                            print("Insufficient balance")
                            print("\n")
                    else:
                        print("Invalid Input")
                        
                elif(ch==2):
                    a=(input("Enter amount to deposit: "))
                    if(a.isdigit()):
                        user.balance=int(user.balance)+int(a)
                        transactions.append((f"${a} has been deposited into your account with user id {u}."))
                        hjk= f"${a} has been deposited into your account with user id {u}."
                        print("Deposit complete,Remaining balance")
                        print(user.balance)
                        print("\n")
                        update_balance_in_excel(users, workbook)
                        update_transaction_history(hjk)
                    else:
                        print("Invalid Input")
                elif(ch==3):
                    receiver_id = int(input("Enter user ID to transfer: "))
                    receiver = next((u for u in users if u.user_id == receiver_id), None)
                    if receiver:
                        amount = int(input("Enter amount to transfer: "))
                        if amount <= user.balance:
                            user.balance -= amount
                            receiver.balance += amount
                            transactions.append((f"${amount} has been transferred to user ID {receiver_id}.", user.balance))
                            hjk= f"${amount} has been transfered from {u} into account with user id {receiver_id}."
                            print("Transfer complete.")
                            print("Your balance:", user.balance)
                            print(f"Receiver's balance: {receiver.balance}\n")
                            update_balance_in_excel(users, workbook)
                            update_transaction_history(hjk)
                        else:
                            print("Insufficient balance")
                            print("\n")
                    else:
                        print("Receiver user ID not found.")
                elif(ch==4):
                    for transaction in transactions:
                        print(transaction)
                        print("Transactions updated please check you sheet")
                    print("\n")
                elif(ch==5):
                    print("Thank You for using the ATM. Have a Nice Day!!!")
                    break
                elif(ch==6):
                    print(user.balance)
            break
    else:
        print("User ID or PIN is incorrect.")

def main():
    print("Welcome to Python ATM\n")
    users, workbook = load_users_from_excel()
    user_id = input("Please Enter your User ID: ")
    pin = pwinput.pwinput(prompt='Enter your PIN: ', mask='*')
    usage(user_id, pin, users, workbook)

if __name__ == "__main__":
    main()
